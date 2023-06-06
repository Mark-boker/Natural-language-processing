# -*- coding: utf-8 -*-

import openpyxl
import pandas as pd
from openpyxl import Workbook
from snownlp import SnowNLP
from snownlp import sentiment


my_list = []  # 按行存放Excel表中数据
wb = openpyxl.load_workbook('评论信息.xlsx')
ws = wb['test_sheet']
maxrows = ws.max_row  # 获取最大行
for i in range(maxrows - 1):
    temp_list = []
    for each in ws.iter_cols(min_row=2):
        temp_list.append(each[i].value)
    my_list.append(temp_list)


filepath = "评论信息.xlsx"
data = pd.read_excel(filepath, names=["名字", "性别", "等级", "评论内容", "点赞数"], usecols=[0, 1, 2, 3, 4])


# 使用默认的词典对模型进行训练
# D:\Software\python\Lib\site-packages\snownlp\sentiment\pos.txt
# D:\Software\python\Lib\site-packages\snownlp\sentiment\neg.txt
sentiment.train(r'pos.txt', r'neg.txt')

sentiment.save('sentiment.bilibili_comment')
pinglun = data['评论内容'].tolist()

senti = [SnowNLP(i).sentiments for i in pinglun]

# 定义列表
newsenti = []
count = 0
for i in senti:
    my_list[count].append(i)
    if (i >= 0.4):
        my_list[count].append("正向")
    else:
        my_list[count].append('负向')
    count = count + 1
names = ["名字", "性别", "等级", "评论内容", "点赞数","SnowNLP评分","情感"]
my_list.insert(0,names)
# 这种方式如果写入的文件不存在，会自动创建文件
workbook = Workbook()
save_file = "D:\文档\PycharmProjects\\network_analysis\大作业\情感输出结果.xlsx"
worksheet = workbook.active
# 每个workbook创建后，默认会存在一个worksheet，对默认的worksheet进行重命名
worksheet.title = "Sheet1"
for row in my_list:
    worksheet.append(row)  # 把每一行append到worksheet中
workbook.save(filename=save_file)  # 不能忘

# 新建评分内容列
# print(newsenti)

# comment = data['评论内容'].tolist()
# print(data['评论内容'].tolist())
# for f in comment:
#     cop = re.compile("[^\u4e00-\u9fa5^a-z^A-Z^0-9]")  # 匹配不是中文、大小写、数字的其他字符
#     nwe_s = cop.sub('', f)  # 将old_s中匹配到的字符替换成空s字符
#     s = SnowNLP(nwe_s)
#     print(s.words)  # 返回分词结果
#     print(s.sentiments)  # 返回该句话的情感得分


